import logging
import os
import sys
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import pyautogui
import requests
import tkinter as tk
import yaml
import keyboard
from cryptography.fernet import Fernet, InvalidToken
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from openpyxl import load_workbook
from retry import retry
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait
from tkinter import messagebox

try:
    import win32com.client
except ImportError:
    win32com = None


@dataclass
class AutomationConfig:
    edge_path: str
    download_folder: str
    template_path: str
    url: str
    username: str
    password: str
    wait_time: Dict[str, int]
    webdriver_path: Optional[str] = None
    contact_sheet_name: Optional[str] = None
    body_sheet_name: Optional[str] = None


class AutomationError(Exception):
    pass


def check_password(password: str) -> bool:
    key = b'yCh_OE7jEoX7S9aUMuk-CCNiJT_GIfb1ZHkLO8b5jbw='
    cipher = b'gAAAAABnwUPuHzq3v84PkAwHhqeiqE2WnXY-2IxdOoZeOHfyeKVTToWfh89_8WQRTPGxJFJ40aoorfQLbb0-3pMRgX-cA2m41g=='
    f = Fernet(key)
    try:
        return password == f.decrypt(cipher).decode()
    except InvalidToken:
        return False
    except Exception as exc:
        raise AutomationError(f"パスワード復号処理でエラーが発生しました: {exc}")


class AutomationScript:
    def __init__(self, config_path: str):
        self.setup_logging()
        self.config = self.load_config(config_path)
        # pyautoguiの設定
        pyautogui.PAUSE = 0.5
        pyautogui.FAILSAFE = True
        # マウスを画面中央に移動
        screen_width, screen_height = pyautogui.size()
        pyautogui.moveTo(screen_width // 2, screen_height // 2)

        self.driver: Optional[webdriver.Edge] = None
        self.browser_wait: Optional[WebDriverWait] = None
        self.running = True
        threading.Thread(target=self._monitor_esc, daemon=True).start()

    def setup_logging(self) -> None:
        log_dir = Path(__file__).parent / "logs"
        log_dir.mkdir(parents=True, exist_ok=True)
        log_file = log_dir / f"automation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s [%(levelname)s] %(message)s',
            handlers=[logging.FileHandler(log_file), logging.StreamHandler()],
        )
        self.logger = logging.getLogger(__name__)

    def load_config(self, path: str) -> AutomationConfig:
        config_path = Path(path)
        if not config_path.exists():
            raise AutomationError(f"設定ファイルが見つかりません: {config_path}")
        with open(config_path, encoding='utf-8') as f:
            data = yaml.safe_load(f) or {}
        if not data.get('edge_path'):
            raise AutomationError("config.yaml に edge_path が設定されていません")
        url = (data.get('url') or "").strip()
        if not url:
            raise AutomationError("config.yaml に url を設定してください")
        username = (data.get('username') or "").strip()
        password = (data.get('password') or "").strip()
        if not username or not password:
            raise AutomationError("config.yaml に username/password を設定してください")
        data['url'] = url
        data['username'] = username
        data['password'] = password
        return AutomationConfig(**data)

    def _monitor_esc(self) -> None:
        while self.running:
            if keyboard.is_pressed('esc'):
                self.logger.warning("ESCキー検知：強制終了")
                self.cleanup(close_browser=True)
                os._exit(0)
            time.sleep(0.3)



    def start_webdriver(self) -> webdriver.Edge:
        options = EdgeOptions()
        options.use_chromium = True
        options.add_argument("--start-maximized")
        download_folder = Path(self.config.download_folder).expanduser().resolve()
        download_folder.mkdir(parents=True, exist_ok=True)
        prefs = {
            "download.default_directory": str(download_folder),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
        }
        options.add_experimental_option("prefs", prefs)
        edge_binary = self.resolve_edge_binary()
        if edge_binary:
            options.binary_location = edge_binary
            self.logger.info(f"Edge実行ファイルを設定しました: {edge_binary}")
        else:
            self.logger.warning("edge_path が設定されていないため既定の Edge を使用します")
        self.logger.info("WebDriver の起動を試みます")
        if self.config.webdriver_path:
            service = webdriver.EdgeService(self.config.webdriver_path)
            driver = webdriver.Edge(service=service, options=options)
        else:
            driver = webdriver.Edge(options=options)
        driver.set_page_load_timeout(60)
        try:
            driver.execute_cdp_cmd(
                "Page.setDownloadBehavior",
                {
                    "behavior": "allow",
                    "downloadPath": str(download_folder),
                    "eventsEnabled": True,
                },
            )
        except Exception as exc:
            self.logger.warning(f"ダウンロード設定の適用に失敗しました: {exc}")
        return driver

    def login(self) -> None:
        if not self.browser_wait:
            raise AutomationError("WebDriverが初期化されていません")
        wait = self.browser_wait
        self.logger.info("ログイン処理を開始します")
        wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='__next']/div/main/div[2]/div[2]/a"))).click()
        time.sleep(self.config.wait_time.get('click', 2))
        user = wait.until(EC.presence_of_element_located((By.ID, "account")))
        user.clear()
        user.send_keys(self.config.username)
        time.sleep(2)
        pwd = wait.until(EC.presence_of_element_located((By.ID, "password")))
        pwd.clear()
        pwd.send_keys(self.config.password)
        time.sleep(2)
        wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='mainContent']/div/div[2]/div[4]/input"))).click()
        time.sleep(max(2, self.config.wait_time.get('browser', 6)))

    def navigate_entries(self) -> None:
        if not self.browser_wait:
            raise AutomationError("WebDriverが初期化されていません")
        self.logger.info("応募者一覧へ遷移します")
        time.sleep(5)
        self.browser_wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='__next']/header/div/nav/ul/li[3]/a"))).click()
        time.sleep(self.config.wait_time.get('browser', 4))

    def filter_entries(self, status_value: str = "01") -> None:
        wait = self.browser_wait
        self.logger.info(f"ステータスを{status_value} に設定して検索します")
        select_element = wait.until(EC.element_to_be_clickable((By.XPATH, "//select[@name='selectionStatus' and @data-select='selectBox']")))
        Select(select_element).select_by_value(status_value)
        self.click_search()
        time.sleep(self.config.wait_time.get('browser', 4))

    def click_search(self) -> None:
        if not self.browser_wait:
            raise AutomationError("WebDriverが未初期化です")
        self.browser_wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='applicationList']/form/div/button"))).click()
        time.sleep(self.config.wait_time.get('click', 2))

    def download_entries(self) -> None:
        if not self.browser_wait:
            raise AutomationError("WebDriverが未初期化です")
        self.logger.info("CSVダウンロードを開始します")
        download_button = self.browser_wait.until(
            EC.presence_of_element_located((By.XPATH, "//button[@data-la='entries_download_btn_click']"))
        )
        self.driver.execute_script("arguments[0].click();", download_button)
        time.sleep(self.config.wait_time.get('browser', 6))

    def get_latest_csv(self) -> str:
        folder = Path(self.config.download_folder).expanduser().resolve()
        self.logger.info(f"ダウンロードフォルダからCSVを探します: {folder}")
        csv_files = list(folder.glob("*.csv"))
        if not csv_files:
            raise AutomationError("CSVファイルが見つかりません")
        latest = max(csv_files, key=lambda f: f.stat().st_ctime)
        self.logger.info(f"最新CSV: {latest}")
        return str(latest)

    def process_data(self, csv_path: str) -> pd.DataFrame:
        self.logger.info(f"CSVを読み込みます: {csv_path}")
        df = pd.read_csv(csv_path)
        df = df.iloc[:, [1, 4, 8, 29, 36]]
        df.columns = ["B", "E", "I", "AD", "AK"]
        df["AD"] = df["AD"].apply(self.clean_branch_name)
        return df

    def clean_branch_name(self, text: Any) -> str:
        if not isinstance(text, str):
            return ""
        parts = text.split('　')
        cleaned = parts[-1] if len(parts) > 1 else text
        return cleaned.strip()

    def build_record_file_stem(self, row: pd.Series) -> str:
        def normalize(value: Any) -> str:
            if isinstance(value, str):
                return value.strip()
            if pd.isna(value):
                return ""
            return str(value).strip()

        parts: List[str] = []
        for key in ("B", "AD", "AK"):
            part = normalize(row.get(key))
            if part:
                parts.append(part)
        combined = "_".join(parts)
        safe_stem = "".join(c for c in combined if c.isalnum() or c in ("_", "-", " "))
        safe_stem = safe_stem.strip()
        return safe_stem or "resume"

    def resolve_template_path(self) -> Path:
        base_dir = Path(__file__).parent
        candidate = Path(self.config.template_path)
        if not candidate.is_absolute():
            candidate = base_dir / candidate
        if not candidate.exists():
            fallback = base_dir / "outlookmail_送付フォーマット.xlsx"
            if fallback.exists():
                self.logger.warning(f"テンプレートファイルが見つからないため {fallback} を使用します")
                candidate = fallback
            else:
                raise AutomationError("メールテンプレートファイルが見つかりません")
        return candidate

    def resolve_edge_binary(self) -> Optional[str]:
        raw_path = (self.config.edge_path or "").replace("%s", "").strip()
        if not raw_path:
            return None
        expanded = Path(raw_path).expanduser()
        return str(expanded) if expanded.exists() else raw_path

    def load_template_data(self) -> None:
        path = self.resolve_template_path()
        try:
            wb = load_workbook(path, data_only=True)
        except Exception as exc:
            raise AutomationError(f"テンプレートファイルの読み込みに失敗しました: {exc}")

        sheet_names = wb.sheetnames
        if not sheet_names:
            raise AutomationError("テンプレートファイルにシートが存在しません")

        contact_sheet_name = self.config.contact_sheet_name or sheet_names[0]
        if contact_sheet_name not in sheet_names:
            raise AutomationError(f"指定された連絡先シートが見つかりません: {contact_sheet_name}")

        try:
            contact_df = pd.read_excel(path, sheet_name=contact_sheet_name, engine="openpyxl")
        except Exception as exc:
            raise AutomationError(f"連絡先リストの読み込みに失敗しました: {exc}")

        contact_df = contact_df.rename(
            columns={
                "拠点名": "branch",
                "担当者": "person",
                "To:": "to",
                "Cc:": "cc",
            }
        )
        contact_df["branch_norm"] = contact_df["branch"].apply(self.clean_branch_name)
        self.contact_df = contact_df

        body_sheet_name = self.config.body_sheet_name
        if body_sheet_name and body_sheet_name not in sheet_names:
            raise AutomationError(f"指定されたメール本文シートが見つかりません: {body_sheet_name}")

        if body_sheet_name:
            template_sheet = wb[body_sheet_name]
        else:
            sheets = wb.worksheets
            template_sheet = sheets[1] if len(sheets) > 1 else sheets[0]

        try:
            self.mail_subject_template = (template_sheet["B1"].value or "").strip()
            raw_body = template_sheet["B2"].value or ""
            self.mail_body_template = str(raw_body).replace("%0a", "\n").strip()
        except Exception as exc:
            raise AutomationError(f"メール本文テンプレートの読み込みに失敗しました: {exc}")

    def find_contact_by_branch(self, branch_name: str) -> Optional[pd.Series]:
        if not hasattr(self, "contact_df"):
            return None
        target = self.clean_branch_name(branch_name)
        if not target:
            return None
        df = self.contact_df
        matches = df[df["branch_norm"] == target]
        if matches.empty:
            return None
        return matches.iloc[0]

    def build_attachments(self, stem: str, allow_png_only: bool = False) -> List[Path]:
        folder = Path(self.config.download_folder).expanduser().resolve()
        pdf_candidates = list(folder.glob(f"{stem}*.pdf"))
        attachments: List[Path] = []
        if pdf_candidates:
            attachments.append(max(pdf_candidates, key=lambda p: p.stat().st_mtime))

        png_candidates = list(folder.glob(f"{stem}*.png"))
        if png_candidates:
            attachments.append(max(png_candidates, key=lambda p: p.stat().st_mtime))

        if not attachments:
            if allow_png_only and not pdf_candidates and png_candidates:
                return attachments
            raise AutomationError(f"添付ファイルが見つかりません: {stem}")
        return attachments

    def confirm_csv_data(self, df: pd.DataFrame) -> bool:
        preview = df[['B', 'E', 'I', 'AD', 'AK']].head(5).to_string(index=False)
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        root.lift()
        root.focus_force()
        result = messagebox.askokcancel(
            "CSVデータ確認",
            f"下記のデータで処理を開始します。\n\n{preview}\n\nOK→続行 / キャンセル→中断"
        )
        root.destroy()
        return result

    def search_and_open(self, full_name: str) -> Optional[str]:
        if not self.browser_wait or not self.driver:
            raise AutomationError("WebDriverが初期化されていません")
        self.logger.info(f"応募者を検索します: {full_name}")
        search_box = self.browser_wait.until(EC.presence_of_element_located((By.NAME, "searchWord")))
        search_box.clear()
        search_box.send_keys(full_name)
        time.sleep(2)
        self.click_search()
        rows = self.browser_wait.until(EC.presence_of_all_elements_located((By.XPATH, "//td[contains(@class, 'styles_tdSelectionStatus')]")))
        if not rows:
            raise AutomationError("検索結果が見つかりません")
        self.logger.info("対応状況セルを開いて詳細画面へ遷移します")
        time.sleep(2)
        try:
            first_row = self.browser_wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "table tbody tr:first-child"))
            )
            first_row.click()
            self.logger.info("最初の行全体をクリックしました")
            time.sleep(2)
        except Exception:
            time.sleep(2)
            self.logger.warning("行全体のクリックに失敗したため、セルを再試行します")
            time.sleep(2)
            first_cell = self.browser_wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "table tbody tr:first-child td:first-child"))
            )
            first_cell.click()
            self.logger.info("セルをクリックして詳細を開きました")
        try:
            resume_button = self.browser_wait.until(
                EC.presence_of_element_located((By.XPATH, "//a[@data-la='entry_detail_resume_btn_click']"))
            )
            pdf_url = resume_button.get_attribute("href")
            if not pdf_url:
                raise AutomationError("レジュメのPDF URLを取得できませんでした")
            time.sleep(2)
            self.logger.info(f"レジュメPDFのURL: {pdf_url[:80]}...")
            time.sleep(2)
            return pdf_url
        except Exception as exc:
            self.logger.warning(f"レジュメボタンが見つからないためスクリーンショットに切り替えます: {exc}")
            time.sleep(2)
            return None

    def download_pdf_from_url(self, pdf_url: str, file_name: str) -> Path:
        if not self.driver:
            raise AutomationError("WebDriverが未初期化です")
        download_folder = Path(self.config.download_folder).expanduser().resolve()
        download_folder.mkdir(parents=True, exist_ok=True)
        safe_stem = "".join(c for c in file_name if c.isalnum() or c in ("_", "-", " ")).strip() or "resume"
        target_name = f"{safe_stem}.pdf"
        self.logger.info(f"PDFダウンロードを開始します url={pdf_url[:80]}...")
        origin_handle = None
        new_window_created = False
        try:
            origin_handle = self.driver.current_window_handle
        except Exception as exc:
            self.logger.debug(f"現在のウィンドウ取得に失敗しました: {exc}")
        try:
            self.driver.switch_to.new_window("tab")
            new_window_created = True
            self.logger.debug("PDFダウンロード用の新規タブを開きました")
        except Exception as exc:
            self.logger.warning(f"新規タブの作成に失敗したため既存タブを使用します: {exc}")
        try:
            self.driver.get(pdf_url)
        except Exception as exc:
            self.logger.warning(f"PDFビューア表示に失敗しましたがダウンロードは継続します: {exc}")

        headers = {}
        try:
            ua = self.driver.execute_script("return navigator.userAgent;")
            headers["User-Agent"] = ua
        except Exception:
            pass
        try:
            cookies = {c["name"]: c["value"] for c in self.driver.get_cookies()}
        except Exception as exc:
            self.logger.warning(f"クッキー取得に失敗しました: {exc}")
            cookies = {}

        target_path = download_folder / target_name
        try:
            with requests.get(pdf_url, headers=headers, cookies=cookies, stream=True, timeout=60) as resp:
                resp.raise_for_status()
                with open(target_path, "wb") as f:
                    for chunk in resp.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
            self.logger.info(f"PDFを保存しました: {target_path}")
        except Exception as exc:
            raise AutomationError(f"PDFダウンロードに失敗しました: {exc}")

        if new_window_created:
            try:
                self.driver.close()
                self.logger.debug("PDFタブを閉じました")
            except Exception as exc:
                self.logger.warning(f"PDFタブのクローズに失敗しました: {exc}")
            if origin_handle:
                try:
                    self.driver.switch_to.window(origin_handle)
                    self.logger.debug("元のタブへ戻りました")
                except Exception as exc:
                    self.logger.warning(f"元のタブへ戻れませんでした: {exc}")
        else:
            try:
                self.driver.back()
                time.sleep(1)
            except Exception as exc:
                self.logger.warning(f"前の画面への戻りに失敗しました: {exc}")
        return target_path

    def capture_screenshot(self, stem: str) -> Path:
        folder = Path(self.config.download_folder).expanduser().resolve()
        folder.mkdir(parents=True, exist_ok=True)
        safe_stem = "".join(c for c in stem if c.isalnum() or c in ("_", "-", " ")).strip() or "screenshot"
        target_path = folder / f"{safe_stem}.png"
        img = pyautogui.screenshot()
        img.save(str(target_path))
        self.logger.info(f"スクリーンショットを保存しました: {target_path}")
        return target_path

    def send_email(self, contact: pd.Series, attachments: List[Path], applicant_email: str = "") -> None:
        if win32com is None:
            raise AutomationError("win32com がインポートできません")

        # To/CC/担当者の値を NaN や空白に対応しつつ安全に取得
        to_raw = contact.get("to", "")
        cc_raw = contact.get("cc", "")
        person_raw = contact.get("person", "")

        to_addr = ""
        if pd.notna(to_raw):
            to_addr = str(to_raw).strip()

        cc_addr = ""
        if pd.notna(cc_raw):
            cc_addr = str(cc_raw).strip()

        person = ""
        if pd.notna(person_raw):
            person = str(person_raw).strip()

        subject = (getattr(self, "mail_subject_template", "") or "").strip()
        body_template = getattr(self, "mail_body_template", "") or ""
        greeting = f"{person} さん" if person else ""

        body_parts: List[str] = []
        if greeting:
            body_parts.append(greeting)
        if applicant_email:
            body_parts.append(f"応募者メール: {applicant_email}")
        if body_template:
            body_parts.append(body_template)
        body = "\n\n".join(body_parts) if body_parts else body_template

        self.logger.info(f"メールを生成します To={to_addr} Cc={cc_addr} 件名={subject}")
        mail = win32com.client.Dispatch("Outlook.Application").CreateItem(0)
        mail.To = to_addr
        if cc_addr:
            mail.CC = cc_addr
        mail.Subject = subject
        mail.Body = body
        for attachment in attachments:
            mail.Attachments.Add(str(attachment))
        mail.Send()
        self.logger.info("メール送信が完了しました")

    def close_overlay(self) -> None:
        if not self.browser_wait:
            return
        try:
            btn = self.browser_wait.until(EC.element_to_be_clickable((By.XPATH, "//img[@data-la='overlay_entry_detail_close_btn_click']")))
            btn.click()
            time.sleep(1)
        except Exception:
            pass

    def update_application_status(self, status_value: str = "04") -> None:
        if not self.browser_wait:
            return
        try:
            select_elem = self.browser_wait.until(
                EC.element_to_be_clickable((By.XPATH, "(//select[@data-select='selectBoxTable'])[1]"))
            )
            Select(select_elem).select_by_value(status_value)
            time.sleep(self.config.wait_time.get('click', 2))
            self.logger.info(f"ステータスを {status_value} に更新しました")
        except Exception as exc:
            self.logger.warning(f"ステータス更新に失敗しました: {exc}")

    def show_dialog(self, message: str, is_error: bool = False) -> None:
        root = tk.Tk()
        root.withdraw()
        if is_error:
            messagebox.showerror("エラー", message)
        else:
            messagebox.showinfo("通知", message)
        root.destroy()

    def cleanup(self, close_browser: bool = True) -> None:
        self.running = False
        if close_browser and self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass

    def run(self) -> None:
        success = False
        try:
            self.driver = self.start_webdriver()
            self.browser_wait = WebDriverWait(self.driver, self.config.wait_time.get('browser', 6) + 10)
            self.driver.get(self.config.url)
            self.login()
            self.navigate_entries()
            self.filter_entries()
            self.download_entries()
            csv_path = self.get_latest_csv()
            df = self.process_data(csv_path)
            if not self.confirm_csv_data(df):
                self.show_dialog("CSV確認で中断しました", is_error=True)
                return
            self.load_template_data()
            for _, row in df.iterrows():
                overlay_closed = False
                try:
                    if int(row["E"]) >= 55:
                        self.logger.info("55歳以上のためスキップ")
                        continue
                    pdf_url = self.search_and_open(row["B"])
                    time.sleep(2)
                    record_stem = self.build_record_file_stem(row)
                    attachments: List[Path] = []
                    pdf_downloaded = False
                    if pdf_url:
                        try:
                            self.download_pdf_from_url(pdf_url, record_stem)
                            pdf_downloaded = True
                        except Exception as exc:
                            self.logger.warning(f"PDFダウンロードに失敗しました: {exc}")
                    if not pdf_downloaded:
                        try:
                            screenshot_path = self.capture_screenshot(record_stem)
                            attachments.append(screenshot_path)
                        except Exception as exc:
                            self.logger.warning(f"スクリーンショット取得に失敗しました: {exc}")
                    time.sleep(2)
                    contact = self.find_contact_by_branch(row["AD"])
                    if contact is None:
                        self.logger.warning(f"支店名に一致する送信先が見つかりません: {row['AD']}")
                        continue
                    try:
                        if pdf_downloaded:
                            attachments = self.build_attachments(record_stem, allow_png_only=True)
                        elif not attachments:
                            attachments = self.build_attachments(record_stem, allow_png_only=True)
                    except Exception as exc:
                        self.logger.warning(f"添付ファイルが見つかりません: {exc}")
                        continue
                    applicant_email = str(row.get("I", "")).strip()
                    if pdf_downloaded:
                        # PDF取得できた場合 → 応募者アドレスは本文に載せない
                        self.send_email(contact, attachments, applicant_email="")
                    else:
                        # スクショのみの場合 → 応募者アドレスを本文に記載
                        self.send_email(contact, attachments, applicant_email=applicant_email)
                    time.sleep(2)
                    self.close_overlay()
                    overlay_closed = True
                    self.update_application_status("04")
                    time.sleep(2)
                finally:
                    if not overlay_closed:
                        self.close_overlay()
                time.sleep(2)
            success = True
        except Exception as exc:
            self.logger.error("処理中に致命的なエラーが発生しました")
            self.logger.exception(exc)
            raise
        finally:
            if success:
                # 正常終了時はブラウザを開いたままにする
                self.cleanup(close_browser=False)
            else:
                # 例外時や強制終了時はブラウザも閉じる
                self.cleanup(close_browser=True)


def main() -> None:
    if len(sys.argv) < 2:
        print("パスワードが必要です")
        sys.exit(1)

    password = sys.argv[1]

    if len(sys.argv) > 2 and sys.argv[2] == "--verify":
        if not check_password(password):
            print("パスワードが違います")
            sys.exit(1)
        sys.exit(0)

    if not check_password(password):
        print("パスワードが違います")
        sys.exit(1)

    config_path = Path(__file__).parent / "config.yaml"
    automation = AutomationScript(str(config_path))
    try:
        automation.run()
    except Exception as exc:
        automation.logger.error("全体処理で致命的なエラーが発生しました")
        automation.logger.exception(exc)
        raise


if __name__ == "__main__":
    main()
